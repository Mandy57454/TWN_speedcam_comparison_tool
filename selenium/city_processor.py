import time
import os
import openpyxl
import re
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException
import pandas as pd

project_path = r"C:\Users\mandy.chang\PycharmProjects\TWN_speedcam_compare"


def wait_for_page_load(driver, timeout=10):
    start_time = time.time()
    while time.time() - start_time < timeout:
        # check page readyState
        ready_state = driver.execute_script("return document.readyState")
        if ready_state == "complete":
            return True
        else:
            time.sleep(0.5)
    return False


def wait_for_element(driver, by, value, timeout=10):
    """
    等待element可見並return element

    :param driver: WebDriver 實例
    :param by: 定位方式（如 By.ID, By.XPATH, By.PARTIAL_LINK_TEXT 等）
    :param value: 定位值
    :param timeout: 超時時間，默認10秒
    :return: 定位到的 WebElement
    """
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((by, value))
        )
        return element
    except Exception as e:
        print(f"無法定位element: {e}")
        return None


def process_taipei(driver, info):
    url = info['url']
    keywords_index_s = info['keywords_index_s']
    try:
        for word, index_s in keywords_index_s.items():
            driver.get(url)
            wait_for_page_load(driver)
            element = driver.find_element(By.PARTIAL_LINK_TEXT, word)
            element.click()
            wait_for_page_load(driver)
            
            # 找到所有 PDF 連結
            pdf_links = driver.find_elements(By.PARTIAL_LINK_TEXT, 'pdf')
            print(f"找到 {len(pdf_links)} 個 PDF 檔案")
            
            # 下載所有 PDF 檔案
            for i, pdf_link in enumerate(pdf_links):
                try:
                    print(f"正在下載第 {i+1} 個 PDF 檔案...")
                    pdf_link.click()
                    time.sleep(2)  # 等待下載開始
                except Exception as e:
                    print(f"下載第 {i+1} 個 PDF 時發生錯誤: {str(e)}")
                    continue
                    
    except Exception as e:
        print(f"處理臺北市資料時發生錯誤: {str(e)}")
        raise



def process_TaoYuan(driver, info):
    url = info['url']
    keywords_index_s = info['keywords_index_s']
    try:
        for word, index_s in keywords_index_s.items():
            driver.get(url)
            wait_for_page_load(driver)
            element = driver.find_element(By.PARTIAL_LINK_TEXT, word)
            element.click()
            wait_for_page_load(driver)
            
            # 找到所有 PDF 連結
            pdf_links = driver.find_elements(By.PARTIAL_LINK_TEXT, 'pdf')
            print(f"找到 {len(pdf_links)} 個 PDF 檔案")
            
            # 下載所有 PDF 檔案
            for i, pdf_link in enumerate(pdf_links):
                try:
                    print(f"正在下載第 {i+1} 個 PDF 檔案...")
                    pdf_link.click()
                    time.sleep(2)  # 等待下載開始
                except Exception as e:
                    print(f"下載第 {i+1} 個 PDF 時發生錯誤: {str(e)}")
                    continue
                    
    except Exception as e:
        print(f"處理桃園市資料時發生錯誤: {str(e)}")
        raise


def process_KeeLung(driver, info):
    url = info['url']
    keywords_index_s = info['keywords_index_s']
    try:
        for word, index_s in keywords_index_s.items():
            driver.get(url)
            wait_for_page_load(driver)
            element = driver.find_element(By.PARTIAL_LINK_TEXT, word)
            element.click()
            pdf_links = driver.find_elements(By.CLASS_NAME, 'fileType')
            for index in index_s:
                pdf_links[int(index)].click()
            time.sleep(5)
    finally:
        driver.quit()


def process_HsinChu(driver, info):
    url = info['url']
    keywords = info['keywords']
    wait = WebDriverWait(driver, 10)
    try:
        for kw in keywords:
            # 1) 每次都先重新载入列表页
            driver.get(url)
            wait_for_page_load(driver)

            # 2) 根据 service_unit_title 文本去定位对应的 <div>，再拿到它外层的 <a>
            title_div = wait.until(EC.presence_of_element_located((
                By.XPATH,
                f"//div[contains(@class,'service_unit_title') and contains(normalize-space(.), '{kw}')]"
            )))
            link = title_div.find_element(By.XPATH, "ancestor::a[1]")

            # 3) 隐藏遮挡层（如果需要）
            driver.execute_script("""
                      const m = document.querySelector('div.modal-165');
                      if(m){ m.style.display='none'; }
                    """)

            # 4) 滚动到中间 & JS click
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", link
            )
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", link)

            # 5) 如果点击后留在新页面，等它加载完再回到列表页
            time.sleep(1)

    finally:
        driver.quit()


def process_newtaipei(driver, info):
    url = info['url']
    fixed_keywords = info['fixed']
    mobile_text = info['mobile']
    workbook = openpyxl.Workbook()
    wait = WebDriverWait(driver, 10)

    for fixed, keyWord in fixed_keywords.items():
        driver.get(url)
        wait_for_page_load(driver)
        element = driver.find_element(By.PARTIAL_LINK_TEXT, fixed)
        element.click()
        # pdf_links = driver.find_elements(By.PARTIAL_LINK_TEXT, keyWord)
        pdf = wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, keyWord)))
        time.sleep(0.3)
        pdf.click()

        # 回到原列表页，否则下一次循环你就一直在 PDF 页面上了
        driver.back()
        time.sleep(1)
####################################################################
    driver.get(url)
    wait_for_page_load(driver)

    element = wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, mobile_text)))
    element.click()
    # 抓表格
    table = wait.until(EC.presence_of_element_located((
        By.XPATH,
        f"//table[caption[contains(normalize-space(.), '{mobile_text}')]]"
    )))
    rows = table.find_elements(By.TAG_NAME, "tr")

    sheet = workbook.create_sheet(title=mobile_text)
    for r, row in enumerate(rows, start=1):
        cells = row.find_elements(By.XPATH, ".//th|.//td")
        for c, cell in enumerate(cells, start=1):
            sheet.cell(row=r, column=c, value=cell.text)
    # 移除默认创建的空白 sheet
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    # 存档
    filename = f"{mobile_text}.xlsx"
    save_path = os.path.join(project_path, "selenium", "city_data", "NewTaipei", filename)
    workbook.save(save_path)
    time.sleep(1)

    driver.quit()


def process_Science_Park(driver, info):
    url = info['url']
    try:
        driver.get(url)
        wait_for_page_load(driver)
        element = driver.find_element(By.XPATH, "//img[@src='/static/images/file/pdf.png']")
        element.click()
        # input("Press Enter to close the browser...")
        time.sleep(2)
    finally:
        driver.quit()


def process_MiaoLi_YunLin(driver, info):
    url = info['url']
    keywords_index_s = info['keywords_index_s']
    try:
        for word, index_s in keywords_index_s.items():
            driver.get(url)
            wait_for_page_load(driver)

            found = False
            while not found:
                try:
                    element = driver.find_element(By.PARTIAL_LINK_TEXT, word)
                    element.click()
                    found = True
                except NoSuchElementException:
                    try:
                        nextpage = driver.find_element(By.PARTIAL_LINK_TEXT, '下一頁')
                        nextpage.click()
                        wait_for_page_load(driver)
                    except NoSuchElementException:
                        print("No more pages and element not found")
                        break

            if found:
                pdf_links = driver.find_elements(By.XPATH, "//img[@src='/Content/img/d06.gif']")
                for index in index_s:
                    try:
                        pdf_links[int(index)].click()
                        time.sleep(3)  # 等待下载或页面加载
                    except IndexError:
                        print(f"Index {index} is out of range for PDF links")
    finally:
        driver.quit()


def process_TaiChung(driver, info):
    url = info['url']
    keywords_index_s = info['keywords_index_s']
    try:
        for word, index_s in keywords_index_s.items():
            driver.get(url)
            wait_for_page_load(driver)
            element = driver.find_element(By.PARTIAL_LINK_TEXT, word)
            element.click()
            pdf_links = driver.find_elements(By.XPATH, "//img[@alt='PDF']")
            for index in index_s:
                pdf_links[int(index)].click()
            time.sleep(2)
    finally:
        driver.quit()


def process_ChiaYi_YiLan(driver, info):
    url = info['url']
    keywords_keywords = info['keywords_keyWords']
    try:
        for keyword, keyWord in keywords_keywords.items():
            driver.get(url)
            wait_for_page_load(driver)
            
            # 直接使用部分匹配尋找關鍵字連結
            element = None
            
            # 方法1: 部分匹配 - 將關鍵字分解成詞組
            keyword_parts = keyword.split()
            for part in keyword_parts:
                if len(part) > 2:  # 只處理長度大於2的詞組
                    try:
                        element = driver.find_element(By.PARTIAL_LINK_TEXT, part)
                        print(f"✅ 找到部分匹配的連結 (關鍵字: '{part}'): {keyword}")
                        break
                    except NoSuchElementException:
                        continue
            
            # 方法2: 如果還是找不到，搜尋所有連結
            if element is None:
                all_links = driver.find_elements(By.TAG_NAME, "a")
                for link in all_links:
                    try:
                        link_text = link.text
                        # 檢查連結文字是否包含關鍵字的任何部分
                        if any(part in link_text for part in keyword_parts if len(part) > 2):
                            element = link
                            print(f"✅ 找到包含關鍵字的連結: {link_text}")
                            break
                    except:
                        continue
            
            if element is None:
                print(f"❌ 無法找到包含關鍵字 '{keyword}' 的連結")
                continue
                
            element.click()
            wait_for_page_load(driver)
            
            # 尋找 PDF 連結 - 直接使用部分匹配
            pdf_links = []
            
            # 方法1: 部分匹配 PDF 關鍵字
            keyWord_parts = keyWord.split()
            for part in keyWord_parts:
                if len(part) > 2:
                    pdf_links = driver.find_elements(By.PARTIAL_LINK_TEXT, part)
                    if pdf_links:
                        print(f"✅ 找到部分匹配的 PDF 連結 (關鍵字: '{part}')")
                        break
            
            # 方法2: 如果還是找不到，搜尋所有 PDF 連結
            if not pdf_links:
                pdf_links = driver.find_elements(By.PARTIAL_LINK_TEXT, 'pdf')
                if not pdf_links:
                    all_links = driver.find_elements(By.TAG_NAME, "a")
                    pdf_links = [link for link in all_links if 'pdf' in link.text.lower()]
            
            print(f"📄 找到 {len(pdf_links)} 個 PDF 連結")
            
            for i, pdf_link in enumerate(pdf_links):
                try:
                    print(f"⬇️ 正在下載第 {i+1} 個 PDF 檔案...")
                    pdf_link.click()
                    time.sleep(2)
                except Exception as e:
                    print(f"❌ 下載第 {i+1} 個 PDF 時發生錯誤: {str(e)}")
                    continue
                    
            time.sleep(3)
    except Exception as e:
        print(f"❌ 處理宜蘭縣資料時發生錯誤: {str(e)}")
        raise


def process_TaiNan(driver, info):
    url = info['url']
    keywords = info['keywords']
    try:
        for keyword in keywords:
            driver.get(url)
            wait_for_page_load(driver)
            element = driver.find_element(By.PARTIAL_LINK_TEXT, keyword)
            element.click()
            time.sleep(2)
    finally:
        driver.quit()


def process_TaiTung(driver, info):
    url = info['url']
    try:
        driver.get(url)
        wait_for_page_load(driver)
        element = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@href,'.pdf')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        time.sleep(0.3)
        element.click()
        time.sleep(0.5)
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[-1])
        time.sleep(40)
    finally:
        driver.quit()


def process_HuaLian_KinMen(driver, info):
    url = info['url']
    try:
        driver.get(url)
        wait_for_page_load(driver)
        pdf_links = driver.find_elements(By.CLASS_NAME, 'pdf')
        # if len(pdf_links) >= 2:
        # 點擊第二個元素（索引從 0 開始，所以第二個元素是索引 1）
        #    pdf_links[1].click()
        # else:
        pdf_links[0].click()
        time.sleep(2)
    finally:
        driver.quit()


def process_Kaohsiung(driver, info):
    url = info['url']
    keywords_keywords = info['keywords_keyWords']

    # 新建一个工作簿，用于合并所有表格
    workbook = openpyxl.Workbook()

    try:
        for keyword, keyWord in keywords_keywords.items():
            driver.get(url)
            wait_for_page_load(driver)
            # element = driver.find_element(By.PARTIAL_LINK_TEXT, keyword)
            element = driver.find_element(By.XPATH, f"//a[text()[contains(., '{keyword}')]]")
            element.click()
            # table = driver.find_element(By.XPATH, "//table[caption[contains(text(), keyWord)]]")
            # 定位表格
            table = driver.find_element(By.XPATH, f"//table[caption[contains(normalize-space(.), '{keyWord}')]]")

            # 如果表头使用 <td> 而非 <th>，则从第一行的 <td> 中提取
            all_rows = table.find_elements(By.TAG_NAME, 'tr')
            # 假设第一条 <tr> 为表头行
            header_row = all_rows[0]
            header_elements = header_row.find_elements(By.TAG_NAME, 'td')
            headers = [h.text.strip() for h in header_elements]
            n_cols = len(headers)

            # 写入新 sheet
            sheet = workbook.create_sheet(title=keyword)
            # 写入表头
            for col_idx, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_idx, value=header)

            # 准备 rowspan 处理
            span_counters = [0] * n_cols
            prev_data = [''] * n_cols
            sheet_row = 2

            # 遍历数据行，从第二行开始
            for row in all_rows[1:]:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if not cells:
                    continue
                row_data = [''] * n_cols
                td_idx = 0
                for col_idx in range(n_cols):
                    if span_counters[col_idx] > 0:
                        row_data[col_idx] = prev_data[col_idx]
                        span_counters[col_idx] -= 1
                    else:
                        cell = cells[td_idx]
                        text = cell.text.strip()
                        row_data[col_idx] = text
                        rs = cell.get_attribute('rowspan')
                        if rs and rs.isdigit() and int(rs) > 1:
                            span_counters[col_idx] = int(rs) - 1
                        td_idx += 1

                # 跳过全空行
                if all(not v for v in row_data):
                    continue

                # 打印调试信息
                # print(f"Keyword={keyword}, 写入行{sheet_row}: {row_data}")

                # 写入 Excel 并尝试将可转为整数的值转换为 int
                for col_idx, val in enumerate(row_data, start=1):
                    # 如果字符串全为数字，则转为 int
                    write_val = int(val) if val.isdigit() else val
                    sheet.cell(row=sheet_row, column=col_idx, value=write_val)

                prev_data = row_data
                sheet_row += 1

            time.sleep(2)

            # 移除默认创建的空白 sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

            # 保存合并文件
            save_dir = os.path.join(project_path, 'selenium', 'city_data', 'Kaohsiung')
            os.makedirs(save_dir, exist_ok=True)
            save_path = os.path.join(save_dir, '高雄市.xlsx')
            workbook.save(save_path)
    finally:
        driver.quit()


def process_ChiaYi_sh(driver, info):
    url = info['url']
    keywords_keywords = info['keywords_keyWords']
    try:
        for keyword, keyWord in keywords_keywords.items():
            driver.get(url)
            wait_for_page_load(driver)
            element = driver.find_element(By.PARTIAL_LINK_TEXT, keyword)
            driver.execute_script("document.querySelector('div.tab').style.display = 'none';")
            element.click()
            wait_for_page_load(driver)
            wait = WebDriverWait(driver, 10)

            download_box = wait.until(EC.presence_of_element_located((By.ID, "download_box")))
            pdf_links = download_box.find_elements(By.CSS_SELECTOR, "a[title$='.pdf']")
            for idx, link in enumerate(pdf_links, start=1):
                print(idx, link.get_attribute("title"), link.get_attribute("href"))
                # 如果要点击，可以：
                driver.execute_script("arguments[0].click();", link)
            time.sleep(10)

    finally:
        driver.quit()


def process_PengHu(driver, info):
    url = info['url']
    try:
        driver.get(url)
        wait_for_page_load(driver)
        wait = WebDriverWait(driver, 10)
        all_data = []
        headers = None
        page = 1

        while True:
            # 等待 table 載入
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'table.table0.table02')))

            # 每次都重新取 row
            rows = table.find_elements(By.CSS_SELECTOR, 'tbody tr')

            # 提取表頭（僅提取一次）
            if headers is None:
                headers = [th.text for th in table.find_elements(By.CSS_SELECTOR, 'thead th')]

            # 提取表格數據
            for row in rows:
                try:
                    cells = row.find_elements(By.CSS_SELECTOR, 'td')
                    row_vals = []
                    for cell in cells:
                        txt = cell.text.strip()
                        # 如果全是数字，转成 int；否则保留原文本
                        if txt.isdigit():
                            row_vals.append(int(txt))
                        else:
                            row_vals.append(txt)
                    all_data.append(row_vals)
                except StaleElementReferenceException:
                    print("Row stale, skip this row")
                    continue

            # 點擊「下一頁」
            try:
                next_page = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "下一頁")))
                next_page.click()
                print(f"已點擊 下一頁，第{page}頁")
                page += 1

                # 點擊後建議等待新的 table reload
                wait.until(EC.staleness_of(table))  # 等舊 table 消失
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'table.table0.table02')))  # 再等新 table 出現

            except TimeoutException:
                print("沒有下一頁，結束爬取")
                break

        # 創建 DataFrame 並保存為 Excel 文件
        import pandas as pd
        filename = f"{'澎湖縣'}.xlsx"
        save_path = os.path.join(project_path, r"selenium\city_data\PengHu", filename)
        df = pd.DataFrame(all_data, columns=headers)
        df.to_excel(save_path, index=False)
        time.sleep(2)
    finally:
        driver.quit()


def process_HsinChu_web(driver, info):
    url = info['url']
    city = info['city']
    try:
        driver.get(url)
        wait_for_page_load(driver)
        main_content = driver.find_element(By.CLASS_NAME, "main-content")
        tables = main_content.find_elements(By.TAG_NAME, "table")
        print(f"抓到 {len(tables)} 個 table！")
        for idx, table in enumerate(tables):
            workbook = openpyxl.Workbook()
            for idx, table in enumerate(tables):
                if idx == 0:
                    sheet = workbook.active
                    sheet.title = f"table{idx + 1}"
                else:
                    sheet = workbook.create_sheet(title=f"table{idx + 1}")
                rows = table.find_elements(By.TAG_NAME, "tr")
                for row_index, row in enumerate(rows, start=1):
                    cells = row.find_elements(By.TAG_NAME, "th")
                    if not cells:
                        cells = row.find_elements(By.TAG_NAME, "td")
                    for cell_index, cell in enumerate(cells, start=1):
                        text = cell.text.strip()
                        try:
                            value = int(text)
                        except ValueError:
                            value = text
                        sheet.cell(row=row_index, column=cell_index, value=value)
            filename = f"{'新竹市'}.xlsx"
            save_path = os.path.join(project_path, r"selenium\city_data", city)
            workbook.save(os.path.join(save_path, filename))
    finally:
        driver.quit()


def process_ChangHua(driver, info):
    url = info['url']
    try:
        driver.get(url)
        wait_for_page_load(driver)
        # driver.find_element(By.CSS_SELECTOR, ".kf_dload-xls").click()
        for el in driver.find_elements(By.CSS_SELECTOR, ".kf_dload-xls"):
            el.click()
            time.sleep(1)
        time.sleep(2)
    finally:
        driver.quit()


def process_NanTou(driver, info):
    url = info['url']
    try:
        driver.get(url)
        wait_for_page_load(driver)
        wait = WebDriverWait(driver, 10)
        all_sheets = []  # 存放每一頁的 DataFrame
        sheet_names = []  # 存放每一頁的名稱


        while True:
            # 1. 取得目前頁碼
            page_div = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.page")))
            page_text = page_div.text
            match = re.search(r"(\d+)/(\d+)", page_text)
            if match:
                current_page = int(match.group(1))
            else:
                print("無法取得目前頁碼")
                break

            # 2. 抓ul-li內容
            ul = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'ul.list.text_le')))
            lis = ul.find_elements(By.TAG_NAME, "li")
            if not lis:
                print("未找到任何資料行，結束")
                break

            # 3. 每一頁都重新抓 header
            spans = lis[0].find_elements(By.TAG_NAME, "span")
            headers = [span.text.strip() for span in spans]
            data_lis = lis[1:]

            # 4. 每筆資料
            page_rows = []
            for li in data_lis:
                spans = li.find_elements(By.TAG_NAME, "span")
                row = [span.text.strip() for span in spans]
                page_rows.append(row)

            # 存成DataFrame
            df = pd.DataFrame(page_rows, columns=headers)
            all_sheets.append(df)
            sheet_names.append(f"第{current_page}頁")

            # 5. 點下一頁，然後比對頁碼
            try:
                next_btn = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "下一頁")))
                next_btn.click()
                print("已點擊 下一頁")
                time.sleep(1)
            except Exception:
                print("沒有下一頁按鈕，結束爬取")
                break

            # 6. 翻頁後再次讀頁碼，比較有沒有變
            page_div = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.page")))
            page_text = page_div.text
            match = re.search(r"(\d+)/(\d+)", page_text)
            if match:
                new_page = int(match.group(1))
            else:
                print("翻頁後無法取得頁碼")
                break

            if new_page == current_page:
                print("頁碼沒變，已經是最後一頁，結束")
                break

        # 存檔（每一頁一個sheet）
        filename = f"{'南投縣'}.xlsx"
        save_path = os.path.join(project_path, r"selenium\city_data\NanTou", filename)
        with pd.ExcelWriter(save_path) as writer:
            for df, name in zip(all_sheets, sheet_names):
                # Excel sheet name 最長31字
                df.to_excel(writer, sheet_name=name[:31], index=False)
        time.sleep(2)
    finally:
        driver.quit()
