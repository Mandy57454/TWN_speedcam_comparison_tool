import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz

project_path = r"C:\Users\mandy.chang\PycharmProjects\TWN_speedcam_compare"


def sort_excel(input_data, output_data):
    df = pd.read_excel(input_data)
    df_sorted = df.sort_values(by=[df.columns[0], df.columns[1], df.columns[2], df.columns[3], df.columns[4]])
    df_sorted.to_excel(output_data, index=False)
    # print(f"表格已排序，并保存到 {output_data}")


def combine_excel():
    folder_path = os.path.join(project_path, r'split_data\output')  # 需要combine的path
    data_combined = os.path.join(project_path, r'compare_data\input\gov_data.xlsx')  # combine後的path

    all_dfs = []
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if filename == 'mio_data.xlsx':
                continue
            file_path = os.path.join(folder_path, filename)
            df = pd.read_excel(file_path)
            # df['Source File'] = filename
            all_dfs.append(df)

    combined_df = pd.concat(all_dfs, ignore_index=True)
    combined_df.to_excel(data_combined, index=False)
    print(f"All files have been successfully merged into {data_combined}")


def split_1_5(input_data, output_data, input_clean):
    df = pd.read_excel(input_data)
    print("列名稱:", df.columns)

    g_column_name = 'camera_type'
    filtered_df = df[df[g_column_name] == 5]
    filtered_df.to_excel(output_data, index=False)
    print(f"篩選完成，結果已保存到 {output_data}")

    df_cleaned = df.drop(filtered_df.index)
    df_cleaned.to_excel(input_clean, index=False)
    print(f"刪除完成，清理後的數據已保存到 {input_clean}")


def compare_data(mio_path, gov_path, output_mio_path, output_gov_path):
    df_mio = pd.read_excel(mio_path)
    df_gov = pd.read_excel(gov_path)
    df_mio['match'] = 0
    # df_mio['match'] = df_mio['match'].astype(int)
    df_gov['match'] = 0
    # df_gov['match'] = df_gov['match'].astype(int)
    df_mio['g_camera_type'] = 0
    df_mio['g_speed'] = 0
    # df_mio['g_camera_type'] = df_mio['g_camera_type'].astype(int)
    df_mio['g_city'] = ''
    df_mio['g_district'] = ''
    df_mio['g_road_1'] = ''
    df_mio['g_road_2'] = ''
    df_mio['g_note'] = ''
    df_mio['g_match'] = ''

    items = {'city', 'district', 'road_1', 'road_2', 'note'}
    g_items = {'g_city', 'g_district', 'g_road_1', 'g_road_2', 'g_note', 'g_match'}
    for item in items:
        df_mio[item] = df_mio[item].astype(str)
        df_gov[item] = df_gov[item].astype(str)

    for g_item in g_items:
        df_mio[g_item] = df_mio[g_item].astype(str)

    df_mio['camera_type'] = df_mio['camera_type'].astype(int)
    df_gov['camera_type'] = df_gov['camera_type'].astype(int)

    def match_records(row, df_gov):
        if row['match'] == 0:
            matches = df_gov[(df_gov['match'] == 0) & (df_gov['city'] == row['city']) &
                             ((df_gov['district'] == row['district']) | (df_gov['district'] == 'nan'))]
            for index, gov_row in matches.iterrows():
                print((gov_row['road_1'], row['road_1']), fuzz.ratio(gov_row['road_1'], row['road_1']))
                print((gov_row['road_2'], row['road_2']), fuzz.ratio(gov_row['road_2'], row['road_2']))
                print((gov_row['note'], row['note']), fuzz.ratio(gov_row['note'], row['note']))
                note_score = fuzz.ratio(gov_row['note'], row['note'])
                note_condition = (gov_row['note'] == 'nan') or (row['note'] == 'nan') or (note_score > 20)
                if fuzz.ratio(gov_row['road_1'], row['road_1']) > 90 and fuzz.ratio(gov_row['road_2'], row['road_2']) > 85 and note_condition:
                    if gov_row['road_2'] == 'nan' and row['road_2'] == 'nan' and note_score < 67:
                        continue
                    df_mio.at[row.name, 'g_city'] = gov_row['city']
                    df_mio.at[row.name, 'g_district'] = gov_row['district']
                    df_mio.at[row.name, 'g_road_1'] = gov_row['road_1']
                    df_mio.at[row.name, 'g_road_2'] = gov_row['road_2']
                    df_mio.at[row.name, 'g_note'] = gov_row['note']
                    df_mio.at[row.name, 'g_match'] = index
                    df_mio.at[row.name, 'g_camera_type'] = gov_row['camera_type']
                    df_mio.at[row.name, 'g_speed'] = gov_row['speed']
                    df_mio.at[row.name, 'match'] = 1
                    df_gov.at[index, 'match'] = 1
                    if fuzz.ratio(gov_row['note'], row['note']) == 100:
                        df_mio.at[row.name, 'match'] = 2
                    return

    df_mio.apply(lambda row: match_records(row, df_gov), axis=1)

    df_mio.to_excel(output_mio_path, index=False, engine='openpyxl')
    df_gov.to_excel(output_gov_path, index=False, engine='openpyxl')

    wb = openpyxl.load_workbook(output_mio_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[6].value != row[10].value and row[10].value != 0:
            row[6].fill = PatternFill(fill_type='solid', start_color='FF0000')
            row[10].fill = PatternFill(fill_type='solid', start_color='FF0000')
        if row[7].value != row[11].value and row[11].value != 0:
            row[7].fill = PatternFill(fill_type='solid', start_color='FFFF00')
            row[11].fill = PatternFill(fill_type='solid', start_color='FFFF00')
    wb.save(output_mio_path)


# sort mio_data
mio_process_sort = os.path.join(project_path, r'split_data\output\mio_data.xlsx')
mio_sorted = os.path.join(project_path, r'compare_data\input\sorted_mio_data.xlsx')
sort_excel(mio_process_sort, mio_sorted)

# combine gov_data and sort gov_data
gov_process_sort = os.path.join(project_path, r'compare_data\input\gov_data.xlsx')
gov_sorted = os.path.join(project_path, r'compare_data\input\sorted_gov_data.xlsx')
# combine_excel()
sort_excel(gov_process_sort, gov_sorted)

# split data
mio_data_5 = os.path.join(project_path, r'compare_data\input\mio_data_5.xlsx')
mio_data_1 = os.path.join(project_path, r'compare_data\input\mio_data_1.xlsx')
split_1_5(mio_sorted, mio_data_5, mio_data_1)

gov_data_5 = os.path.join(project_path, r'compare_data\input\gov_data_5.xlsx')
gov_data_1 = os.path.join(project_path, r'compare_data\input\gov_data_1.xlsx')
split_1_5(gov_sorted, gov_data_5, gov_data_1)

# compare path
input_path = os.path.join(project_path, r"compare_data\input")
output_path = os.path.join(project_path, r"compare_data\output")

# Compare mio_data_1.xlsx and gov_data_1.xlsx
mio_path = os.path.join(input_path, 'mio_data_1.xlsx')
gov_path = os.path.join(input_path, 'gov_data_1.xlsx')
output_mio_path = os.path.join(output_path, 'mio_out_data_1.xlsx')
output_gov_path = os.path.join(output_path, 'gov_out_data_1.xlsx')
compare_data(mio_path, gov_path, output_mio_path, output_gov_path)

# Compare mio_data_5.xlsx and gov_data_5.xlsx
mio_path = os.path.join(input_path, 'mio_data_5.xlsx')
gov_path = os.path.join(input_path, 'gov_data_5.xlsx')
output_mio_path = os.path.join(output_path, 'mio_out_data_5.xlsx')
output_gov_path = os.path.join(output_path, 'gov_out_data_5.xlsx')
compare_data(mio_path, gov_path, output_mio_path, output_gov_path)
