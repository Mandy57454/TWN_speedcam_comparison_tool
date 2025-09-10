from openpyxl import load_workbook
import os

need_process_file = "臺北市政府警察局固定式違規照相設備及區間測速裝置設置地點一覽表"

# 輸入檔案與工作表名稱
project_path = r"C:\Users\mandy.chang\PycharmProjects\TWN_speedcam_compare"
input_path = os.path.join(project_path, r"converter\city_data\taipei",
                          f"{need_process_file}.xlsx")
output_path = os.path.join(project_path, r"converter\city_data\taipei", f"{need_process_file}.xlsx")
# sheet_name = "Sheet1"  # 或使用 wb.active 取得第一個工作表


def process_taipei_data():
    # 載入 Excel 檔案
    wb = load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]

    # 設定處理的範圍：從第2列開始（假設第1列是表頭）
    for row in range(2, ws.max_row + 1):
        cell_c = ws.cell(row=row, column=3)  # A欄
        cell_d = ws.cell(row=row, column=4)  # B欄

        if cell_c.value and cell_d.value:
            # 將字串合併，寫入 A欄
            merged_text = str(cell_c.value) + str(cell_d.value)
            cell_c.value = merged_text

            # 合併 A與B欄儲存格
            cell_range = f"C{row}:D{row}"
            ws.merge_cells(cell_range)

    # 儲存為新檔案
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb.save(output_path)
    print(f"儲存成功：{output_path}")
