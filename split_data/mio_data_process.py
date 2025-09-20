import pandas as pd
import os
import openpyxl

# 動態取得專案路徑
project_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

need_process_path = os.path.join(project_path, r"split_data\TWN Speed cam update.xlsx")
output_path = os.path.join(project_path, r"split_data\input")


def tec(tec_sheet, data_rows):
    tec_data = []
    for row in tec_sheet.iter_rows(min_row=2, min_col=2, max_col=10, values_only=True):
        tec_data.append(row)
    # 遍历列表并创建数据框的行
    for i in tec_data:
        # print(i)
        tec_row_dict = {
            'city': i[0],
            'address': i[3],
            'camera_type': i[8],
            'speed': i[7],
            'type': 'tec'
        }
        data_rows.append(tec_row_dict)


def fixed(fixed_sheet, data_rows):
    fixed_data = []
    for row in fixed_sheet.iter_rows(min_row=2, min_col=5, max_col=16, values_only=True):
        fixed_data.append(row)
    for i in fixed_data:
        fixed_row_dict = {
            'city': i[9],
            'district': i[10],
            'address': i[11],
            'camera_type': i[1],
            'speed': i[0],
            'type': 'fixed'
        }
        data_rows.append(fixed_row_dict)


def mobile(mobile_sheet, data_rows):
    mobile_data = []
    for row in mobile_sheet.iter_rows(min_row=2, min_col=5, max_col=16, values_only=True):
        mobile_data.append(row)
    for i in mobile_data:
        mobile_row_dict = {
            'city': i[10],
            'address': i[11],
            'camera_type': i[1],
            'speed': i[0],
            'type': 'mobile'
        }
        data_rows.append(mobile_row_dict)


def average(average_sheet, data_rows):
    average_data = []
    for row in average_sheet.iter_rows(min_row=2, min_col=5, max_col=15, values_only=True):
        average_data.append(row)
    for i in average_data:
        average_row_dict = {
            'city': i[10],
            'address': i[8],
            'camera_type': i[1],
            'speed': i[0],
            'type': 'average'
        }
        data_rows.append(average_row_dict)


def combine(combine_sheet, data_rows):
    combine_data = []
    for row in combine_sheet.iter_rows(min_row=2, min_col=5, max_col=14, values_only=True):
        combine_data.append(row)
    for i in combine_data:
        combine_row_dict = {
            'city': i[7],
            'district': i[8],
            'address': i[9],
            'camera_type': i[1],
            'speed': i[0],
            'type': 'combine'
        }
        data_rows.append(combine_row_dict)


def motor(motor_sheet, data_rows):
    motor_data = []
    for row in motor_sheet.iter_rows(min_row=2, min_col=2, max_col=8, values_only=True):
        motor_data.append(row)
    for i in motor_data:
        motor_row_dict = {
            'city': i[0],
            'address': i[1],
            'camera_type': i[6],
            'speed': i[5],
            'type': 'motor'
        }
        data_rows.append(motor_row_dict)


def mobile_del(mobile_del_sheet, data_rows):
    mobile_del_data = []
    for row in mobile_del_sheet.iter_rows(min_row=2, min_col=5, max_col=16, values_only=True):
        mobile_del_data.append(row)
    for i in mobile_del_data:
        mobile_del_row_dict = {
            'city': i[10],
            'address': i[11],
            'camera_type': 5,
            'speed': i[0],
            'type': 'mobile_del'
        }
        data_rows.append(mobile_del_row_dict)


def fixed_del(fixed_del_sheet, data_rows):
    fixed_del_data = []
    for row in fixed_del_sheet.iter_rows(min_row=2, min_col=5, max_col=16, values_only=True):
        fixed_del_data.append(row)
    for i in fixed_del_data:
        fixed_del_row_dict = {
            'city': i[9],
            'district': i[10],
            'address': i[11],
            'camera_type': 0,
            'speed': i[0],
            'type': 'fixed_del'
        }
        data_rows.append(fixed_del_row_dict)


def mio_data_process():
    process_data = openpyxl.load_workbook(need_process_path)
    tec_sheet = process_data["Taiwan科技執法"]
    fixed_sheet = process_data["Taiwan固定式測速"]
    mobile_sheet = process_data["Taiwan移動式"]
    average_sheet = process_data["Taiwan區間測速"]
    combine_sheet = process_data["固定式合併科技執法"]
    motor_sheet = process_data["機車"]
    mobile_del_sheet = process_data["移動式not found or 刪除"]
    fixed_del_sheet = process_data["固定式not found or 刪除"]

    # 初始化空的DataFrame
    df_mio = pd.DataFrame()
    df_mio['city'] = None
    df_mio['district'] = None
    df_mio['road_1'] = None
    df_mio['road_2'] = None
    df_mio['note'] = None
    df_mio['address'] = None
    df_mio['camera_type'] = None
    df_mio['speed'] = None
    df_mio['type'] = None
    data_rows = []

    tec(tec_sheet, data_rows)
    fixed(fixed_sheet, data_rows)
    mobile(mobile_sheet, data_rows)
    average(average_sheet, data_rows)
    combine(combine_sheet, data_rows)
    motor(motor_sheet, data_rows)
    mobile_del(mobile_del_sheet, data_rows)
    fixed_del(fixed_del_sheet, data_rows)

    # 将数据行列表转换为数据框
    df_mio = pd.DataFrame(data_rows,
                          columns=['city', 'district', 'road_1', 'road_2', 'note', 'address', 'camera_type', 'speed',
                                   'type'])
    df_mio = df_mio[df_mio['camera_type'] != 9128]
    df_mio = df_mio[df_mio['address'].notna() & (df_mio['address'] != '')]
    output_mio_path = os.path.join(output_path, 'mio_data.xlsx')
    df_mio.to_excel(output_mio_path, index=False, engine='openpyxl')


mio_data_process()
